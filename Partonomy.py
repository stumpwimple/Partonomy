import random
from collections import OrderedDict

import pickle
import os
os.system('cls')
import copy

import PySimpleGUI as sg
from openpyxl import load_workbook

min_dc = 1
max_dc = 12

unique_game_id = 0
unique_build_id = 0
build_trait_parent_id = "0"
game_trait_parent_id = 0

list_of = {}
categories_of_list = {}

the_aggregate = OrderedDict()


# build_aggregate = OrderedDict()

class Trait:
    """Trait Class - A class that displays a trait ui, with a number of other ui elements meant for running a pprpg"""

    def __init__(self, trait_name, trait_id, parent_id, children_id=[], rand_table="", trait_type="untyped",
                 primary_trait=False, trait_unit=""):
        self.trait_name = trait_name
        self.trait_id = trait_id
        self.parent_id = parent_id
        self.children_id = children_id
        self.trait_type = trait_type
        self.rand_table = rand_table
        self.trait_result = sr(random.choice(list_of[self.rand_table]))
        self.trait_unit = trait_unit
        self.is_discovered = False
        self.primary_trait = primary_trait

        self.journal_entries = {}

    def __str__(self):
        trait_str = str(self.trait_name + ":")
        # print only the left 60 characters of value
        trait_str += str(self.trait_result) + " " + str(self.trait_unit)
        trait_str += str(self.proficiency)
        return trait_str

    def generate(self, trait_name=0):
        pass

    def reroll_trait(self):
        # self.value = eval(self.formula)
        self.proficiency = random.randint(min_dc, max_dc)

class Build_Logic:
    def __init__(self, type="NoLogic", if_else=False,if_test="",if_test_string="", if_true_trait="", if_else_trait="",
                 chance_percent="",if_condition="",quantity_count=1, case_case="",case_test="",case_roll=""):
        self.type = type
        if self.type == "NoLogic":
            self.button_text="LOGIC"
        elif self.type == "If-Then":
            self.if_else = if_else
            self.if_test = if_test
            self.if_condition = if_condition
            self.if_test_string = if_test_string
            self.if_true_trait = if_true_trait
            self.if_else_trait = if_else_trait
            self.button_text="IF THEN"
        elif self.type == "Chance":
            self.chance_percent = chance_percent
            self.button_text="CHANCE " + str(chance_percent)
        elif self.type == "Quantity":
            self.quantity_count = quantity_count
            self.button_text="QUANTITY " + str(quantity_count)
        elif self.type == "Case":
            self.case_case = case_case
            self.case_test = case_test
            self.case_roll = case_roll
            self.button_text="CASE"
        else:
            self.button_text="WTF"



    def __str__(self):
        print_str=""
        if self.type == "NoLogic":
            print_str = "No Logic"
        elif self.type == "If-Then":
            print_str = "If-Then"
            print_str += str(self.if_else)
            print_str += str(self.if_test)
            print_str += str(self.if_condition)
            print_str += str(self.if_test_string)
            print_str += str(self.if_true_trait)
            print_str += str(self.if_else_trait)
        elif self.type == "Chance":
            print_str = "Chance"
            print_str += str(self.chance_percent)
        elif self.type == "Quantity":
            print_str = "Quantity"
            print_str += str(self.quantity_count)
        elif self.type == "Case":
            print_str = "Case"
            print_str += str(self.case_case)+"||"+str(self.case_test)+"||"+str(self.case_roll)
        return print_str

    print("in Build_Logic")

class Build_Trait:

    def __init__(self, trait_id, trait_name="", build_parent_id="0", game_parent_id="0",children_id=[], trait_logic=Build_Logic(),
                 rand_table="Error", trait_type="untyped", primary_trait=False, trait_unit=""):
        self.trait_name = trait_name
        self.trait_id = trait_id
        self.build_parent_id = build_parent_id
        self.game_parent_id = game_parent_id
        self.child_id_list = children_id
        self.trait_type = trait_type
        self.trait_logic = trait_logic
        self.rand_table = rand_table
        self.trait_unit = trait_unit
        self.primary_trait = primary_trait

        # These Traits pertain to gameplay.
        self.trait_result = sr(random.choice(list_of[self.rand_table]))
        self.proficiency_result = random.randint(min_dc, max_dc)
        self.is_discovered = False
        self.journal_entries = {}

    def __str__(self):
        trait_str = str(self.trait_id) + "-" + self.trait_name + ", Result:"
        trait_str += str(self.trait_result) + ", with proficiency:" + str(self.proficiency_result)
        trait_str += ", PARENTS:(b-" + str(self.build_parent_id) + " / g-" + str(self.game_parent_id)
        trait_str += "), CHILDREN: " + str(self.child_id_list)
        return trait_str

    def get_trait_info(self):
        trait_info =  str(self.trait_id) + ":" + self.trait_name
        return trait_info

    def reroll_all(self):
        self.trait_result = sr(random.choice(list_of[self.rand_table]))
        self.proficiency_result = random.randint(min_dc, max_dc)

    def reroll_trait(self):
        self.trait_result = sr(random.choice(list_of[self.rand_table]))

    def reroll_prof(self):
        self.proficiency_result = random.randint(min_dc, max_dc)

    def generate(self, build_agg, new_game_parent_id=0):
        global unique_game_id, game_aggregate
        chance_fail=False

        if self.trait_logic.type == "NoLogic":
            game_aggregate[unique_game_id] = copy.copy(self)
            game_aggregate[unique_game_id].child_id_list = []
            game_aggregate[unique_game_id].game_parent_id=new_game_parent_id
            game_aggregate[unique_game_id].reroll_all()
            game_aggregate[unique_game_id].trait_id=unique_game_id
            print(game_aggregate[unique_game_id])

        elif self.trait_logic.type == "Chance":
            if self.trait_logic.chance_percent > random.randint(1,100):
                chance_fail = False
                game_aggregate[unique_game_id]=copy.copy(self)
                game_aggregate[unique_game_id].child_id_list = []
                game_aggregate[unique_game_id].game_parent_id=new_game_parent_id
                game_aggregate[unique_game_id].reroll_all()
                game_aggregate[unique_game_id].trait_id=unique_game_id
                print(game_aggregate[unique_game_id])
            else:
                chance_fail = True
                print("Failed chance", game_aggregate[new_game_parent_id].child_id_list)
                game_aggregate[new_game_parent_id].child_id_list.remove(unique_game_id)
                unique_game_id -=1

        elif self.trait_logic.type == "Quantity":
            temp_trait = copy.deepcopy(self)
            temp_trait.trait_logic.type="Clone"

            game_aggregate[unique_game_id]=copy.copy(self)
            game_aggregate[unique_game_id].child_id_list = []
            game_aggregate[unique_game_id].game_parent_id=new_game_parent_id
            game_aggregate[unique_game_id].reroll_all()
            game_aggregate[unique_game_id].trait_id=unique_game_id
            prime_game_id=copy.copy(unique_game_id)
            print(game_aggregate[unique_game_id])

            for a in range(1,self.trait_logic.quantity_count):
                unique_game_id +=1
                game_aggregate[new_game_parent_id].child_id_list.append(unique_game_id)
                temp_trait.generate(build_agg, new_game_parent_id)
                #game_aggregate[unique_game_id].game_parent_id = prime_game_parent

        elif self.trait_logic.type == "Case":
            #THIS IS CURRENTLY USING NO LOGIC
            #print("In case, CURRENTLY USING NO LOGIC")
            game_aggregate[unique_game_id]=copy.copy(self)
            game_aggregate[unique_game_id].child_id_list = []
            game_aggregate[unique_game_id].game_parent_id=new_game_parent_id
            game_aggregate[unique_game_id].reroll_all()
            game_aggregate[unique_game_id].trait_id=unique_game_id
            print(game_aggregate[unique_game_id])

        elif self.trait_logic.type == "If-Then":
            #THIS IS CURRENTLY USING NO LOGIC
            #print("In If-Then, CURRENTLY USING NO LOGIC")
            game_aggregate[unique_game_id]=copy.copy(self)
            game_aggregate[unique_game_id].child_id_list = []
            game_aggregate[unique_game_id].game_parent_id=new_game_parent_id
            game_aggregate[unique_game_id].reroll_all()
            game_aggregate[unique_game_id].trait_id=unique_game_id
            print(game_aggregate[unique_game_id])

        elif self.trait_logic.type == "Clone":
            game_aggregate[unique_game_id] = copy.copy(self)
            game_aggregate[unique_game_id].child_id_list = []
            game_aggregate[unique_game_id].game_parent_id=new_game_parent_id
            game_aggregate[unique_game_id].reroll_all()
            game_aggregate[unique_game_id].trait_id = unique_game_id
            print("CLONE:",game_aggregate[unique_game_id])

        else:
            print("Should have logic, but it is not implemented, reroll_all() instead")

        if self.child_id_list != [] and not chance_fail:
            if self.trait_logic.type == "Quantity":
                curr_id=copy.copy(prime_game_id)
            else:
                curr_id=copy.copy(unique_game_id)

            for ch_id in self.child_id_list:
                temp_trait = copy.copy(build_agg[str(ch_id)])
                unique_game_id +=1
                game_aggregate[curr_id].child_id_list.append(unique_game_id)
                temp_trait.generate(build_agg, curr_id)




        # for key in game_aggregate:
        #     print(game_aggregate[key])

def import_random_lists_from_file():
    """ imports excel workbook, creates random generation lists, and list of random list categories.

    Also collects a list of categories this list could belong to and stores it in a list.
    """
    wb = load_workbook(filename='Random Lists.xlsx')
    ws = wb["SHORT_LISTS"]  ### CURRENTLY SET TO SHORT_LISTS FOR FASTER LOAD DEFAULT SHOULD BE "LISTS"

    for col in ws.iter_cols():
        this_list = []
        list_categories = []
        for cell in col:
            # stores list_name for this list of random items
            if cell.row == 1 and cell.value:
                list_name = cell.value
            # stores categories of list and stores them. Excel file uses rows 2-10 as categories
            elif (1 < cell.row <= 10) and cell.value:
                list_categories.append(cell.value)
            # stores items in list of random items
            elif cell.value:
                this_list.append(cell.value)
            else:
                break
        list_of[list_name] = this_list
        categories_of_list[list_name] = list_categories

# SUB-RESOLVE RESULTS WITH NESTED RANDOM LISTS
def sr(this_str):
    this_str=str(this_str)
    """Sub-Resolves any nested Random Lists"""
    while "[" in this_str:
        nested_list = this_str[this_str.find("[") + 1:this_str.find("]")]
        this_result = random.choice(list_of[nested_list])
        try:
            this_result = str(this_result)
        except ValueError:
            pass
        this_str = this_str.replace("[" + nested_list + "]", "|" + this_result + "|")
    return this_str

def build_trait_layout():
    global unique_build_id
    # print ("unique_build_id = ",unique_build_id)
    if unique_build_id == 0:
        button_element = sg.Button('SAVE BUILD', key=("SAVE_BUILD"))
        logic_disabled = True
    else:
        button_element = sg.Button('EXPAND', key=("EXPAND_TRAIT" + str(unique_build_id)), disabled=True)
        logic_disabled = False

    layout = [[sg.Text(unique_build_id, size=(5, 1)),
               sg.Input(unique_build_id, size=(5, 1), k="BUILD_ID" + str(unique_build_id), disabled=True),
               sg.Input(size=(25, 1), key=("TRAIT_NAME" + str(unique_build_id))),
               sg.Button('LOGIC', key=("LOGIC" + str(unique_build_id)), disabled=True, size=(10, 1)),
               sg.Combo(values=[*list_of], size=(25, 1), key=("TRAIT_TABLE" + str(unique_build_id))),
               sg.Input(size=(15, 1), key=("TRAIT_UNIT" + str(unique_build_id))),
               sg.Checkbox('', k='is_primary' + str(unique_build_id)), button_element]]
    return layout

def game_trait_layout():
    global unique_game_id
    if unique_game_id == 0:
        button_element = sg.Button('SAVE GAME', key=("SAVE_GAME"))
        logic_disabled = True
    else:
        button_element = sg.Button('EXPAND', key=("EXPAND_GAME_TRAIT" + str(unique_game_id)), disabled=True)
        logic_disabled = False

    layout = [[sg.Text(unique_game_id, size=(5, 1)),
               sg.Input(unique_game_id, size=(5, 1), k="GAME_ID" + str(unique_game_id), disabled=True),
               sg.Input(size=(5, 1), k="GAME_PROFICIENCY_RESULT" + str(unique_game_id)),
               sg.Button("R", size=(5, 1), k="GAME_ROLL_PROFICIENCY" + str(unique_game_id)),
               sg.Input(size=(20, 1), key=("GAME_TRAIT_NAME" + str(unique_game_id))),
               sg.Input(size=(40, 1), key=("GAME_TRAIT_RESULT" + str(unique_game_id))),
               sg.Button("Roll", size=(5, 1), k="GAME_REROLL_TRAIT" + str(unique_game_id)),
               sg.Button("G", size=(5, 1), k="GAME_SEARCH" + str(unique_game_id)),
               sg.Button("J", size=(5, 1), k="GAME_JOURNAL" + str(unique_game_id)),
               button_element]]
    return layout

def build_layout(build_aggregate):
    global unique_build_id
    global screen_w
    global screen_h

    layout = [[sg.Button('TO PARENT', key=("GO_TO_PARENT")), sg.Text("PARENTS NAME HERE", key="PARENTS_PARENT"),
               sg.Text("", size=(60, 1)), sg.Button('GENERATE AGGREGATE', key=("GENERATE_AGGREGATE"))]]

    layout += [[sg.Text("lay_id", size=(5, 1)),
                sg.Text("bld_id", size=(5, 1)),
                sg.Text("LOGIC", size=(20, 1)),
                sg.Text("LOGIC", size=(10, 1)),
                sg.Text("Trait Roll", size=(23, 1)),
                sg.Text("Units", size=(13, 1)),
                sg.Text("Prim?")]]
    layout += build_trait_layout()
    unique_build_id += 1

    layout += [[sg.Button("LOAD AGG", k = "LOAD_BUILD", size=(13, 1)), sg.Text("Sub-Trait", size=(22, 1))]]

    trait_layout = [[]]

    for i in range(1, 26):
        new_trait_layout = build_trait_layout()
        unique_build_id += 1
        trait_layout += new_trait_layout

    layout += [[sg.Column(trait_layout, size=(int(screen_w / 2), int(screen_h * .70)))]]

    return layout

def game_layout():
    global unique_game_id
    global screen_w
    global screen_h

    layout = [[sg.Button('TO PARENT', key=("GO_TO_GAME_PARENT")),
               sg.Text("GAME PARENTS NAME HERE", key="GAME_PARENTS_PARENT"), sg.Text("", size=(60, 1)),
               sg.Button('REGENERATE AGGREGATE', key=("GAME_REGENERATE_AGGREGATE"))]]

    layout += [[sg.Text("lay_id", size=(5, 1)),
                sg.Text("gam_id", size=(5, 1)),
                sg.Text("Prof", size=(5, 1)),
                sg.Text("R.Prf", size=(5, 1)),
                sg.Text("Trait Name", size=(20, 1)),
                sg.Text("Trait Result", size=(40, 1))]]
    layout += game_trait_layout()
    unique_game_id += 1

    layout += [[sg.Text("", size=(13, 1)), sg.Text("Sub-Trait", size=(22, 1))]]

    trait_layout = [[]]

    for i in range(1, 26):
        new_trait_layout = game_trait_layout()
        unique_game_id += 1
        trait_layout += new_trait_layout

    layout += [[sg.Column(trait_layout, size=(int(screen_w / 2), int(screen_h * .70)))]]

    return layout

def logic_tabbed_window():
    if_then_layout = [[sg.Text("If Then Logic Settings"), sg.Text("",size=(20,1)), sg.Checkbox("Else?", key="if_else")],
                      [sg.Text("If: "), sg.Combo(values=aggregate_list, size=(30, 1), key="if_test"), #Need to be changed to list of existing traits in aggregate
                       sg.Combo(values=('<','>','=','!='),size=(5, 1), key="if_condition"),
                       sg.Combo(values="",size=(30, 1), key="if_test_string")],
                       [],
                       [sg.Text("If True Result = "),sg.Combo(values=[*list_of], key="if_true_trait")],
                       [sg.Text("If False Result = "),sg.Combo(values=[*list_of], key="if_else_trait")]]

    quantity_layout = [[sg.Text("Quantity Logic Settings")],
                       [sg.Text("Num of repeated Traits (2-9)"), sg.Input(key="quantity_count", size =(5,1))]]

    chance_layout = [[sg.Text("Chance Logic Settings")],
                     [sg.Text("Chance of Trait: "), sg.Input(key="chance_percent", size =(5,1))]]


    case_layout = [[sg.Text("Case Logic Settings")]]
    case_layout += [[sg.Text("CASE 1:",size=(20,1)), sg.Combo(values=aggregate_list, size=(30, 1), key="case_1"), sg.Text("="),
                     sg.Combo(values=('Select a Case'), size=(30, 1), key="case_test_1"),
                     sg.Text("   Trait Roll"), sg.Combo(values=[*list_of], size=(30, 1), key="case_roll_1")]]

    for i in range(2,6):
        case_layout+= [[sg.Text("",size=(20,1)),sg.Text("",size=(30,1)),
                        sg.Text("="),sg.Combo(values=('possible roll results','more','+more'),size=(30,1), key="case_test_"+str(i)), sg.Text("   Trait Roll"), sg.Combo(values=[*list_of], size=(30, 1), key="case_roll_"+str(i))]]

    exploding_layout = [[sg.Text("Exploding Logic Settings")]]
    inherit_layout = [[sg.Text("Inherit Logic Settings")]]

    inherit_layout += [[]]

    layout = [[sg.TabGroup(
        [[sg.Tab('NoLogic', [[sg.Text("Logic set to default or NoLogic")]]),
          sg.Tab('If-Then', if_then_layout),
          sg.Tab('Quantity', quantity_layout),
          sg.Tab('Chance', chance_layout),
          sg.Tab('Case', case_layout),
          #sg.Tab('Exploding', exploding_layout),
          #sg.Tab('Inherit', inherit_layout)
                ]],k="Logic_Tabs",)],
              [sg.Text('',size=(40,1)), sg.Button('CANCEL', k='Logic_cancel'),
               sg.Button('OK', k='Logic_ok')]]

    return sg.Window("Select Logic", layout, modal=True)

def initialize_test_agg(build_aggregate):
    global unique_build_id
    unique_build_id = 0
    build_aggregate[str(unique_build_id)] = Build_Trait(str(unique_build_id), trait_name="Solar",
                                                        build_parent_id="0",
                                                        rand_table="SunType",
                                                        primary_trait=True)
    build_aggregate[str(unique_build_id)].child_id_list = []

    # for key in build_aggregate:
    #     print (build_aggregate[key])

    unique_build_id += 1

    build_aggregate[str(unique_build_id)] = Build_Trait(str(unique_build_id), trait_name="SpaceStoryHook?",
                                                        build_parent_id="0",
                                                        rand_table="SpaceStoryHook",
                                                        primary_trait=True)
    build_aggregate[str(unique_build_id)].child_id_list = []

    build_aggregate[str(build_aggregate[str(unique_build_id)].build_parent_id)].child_id_list.append(str(unique_build_id))
    unique_build_id += 1

    for planet in range(1,4):
        build_aggregate[str(unique_build_id)] = Build_Trait(str(unique_build_id), trait_name="Planet " + str(planet),
                                                            build_parent_id="0",
                                                            rand_table="PlanetClass",
                                                            primary_trait=True)
        build_aggregate[str(unique_build_id)].child_id_list = []
        build_aggregate[str(build_aggregate[str(unique_build_id)].build_parent_id)].child_id_list.append(str(unique_build_id))
        planet_id = build_aggregate[str(unique_build_id)].trait_id
        unique_build_id += 1

        build_aggregate[str(unique_build_id)] = Build_Trait(str(unique_build_id), trait_name="CivilizationStoryHook?",
                                                            build_parent_id=str(planet_id),
                                                            rand_table="CivilizationStoryHook",
                                                            primary_trait=True)
        build_aggregate[str(unique_build_id)].child_id_list = []
        build_aggregate[str(build_aggregate[str(unique_build_id)].build_parent_id)].child_id_list.append(str(unique_build_id))
        unique_build_id += 1

        num_regions = random.randint(4,12)
        for region in range(1,num_regions):
            build_aggregate[str(unique_build_id)] = Build_Trait(str(unique_build_id), trait_name="Region " + str(region),
                                                                build_parent_id=str(planet_id),
                                                                rand_table="RockyWaterTopography",
                                                                primary_trait=True)
            build_aggregate[str(unique_build_id)].child_id_list = []
            build_aggregate[str(build_aggregate[str(unique_build_id)].build_parent_id)].child_id_list.append(str(unique_build_id))
            region_id = build_aggregate[str(unique_build_id)].trait_id
            unique_build_id += 1

            build_aggregate[str(unique_build_id)] = Build_Trait(str(unique_build_id), trait_name="RegionStoryHook?",
                                                                build_parent_id=str(region_id),
                                                                rand_table="RegionStoryHook",
                                                                primary_trait=True)
            build_aggregate[str(unique_build_id)].child_id_list = []
            build_aggregate[str(build_aggregate[str(unique_build_id)].build_parent_id)].child_id_list.append(str(unique_build_id))
            unique_build_id += 1

def get_next_game_id():
    global game_aggregate, unique_game_id
    next_id = (int(next(reversed(game_aggregate))) + 1)
    unique_game_id += 1
    return next_id

def update_ui(window,new_parent_id,build_aggregate):
    window['BUILD_ID0'].update(value=new_parent_id)
    window['TRAIT_NAME0'].update(value=build_aggregate[new_parent_id].trait_name)
    window['TRAIT_TABLE0'].update(value=build_aggregate[new_parent_id].rand_table)
    window['TRAIT_UNIT0'].update(value=build_aggregate[new_parent_id].trait_unit)
    window['is_primary0'].update(value=build_aggregate[new_parent_id].primary_trait)
    window['LOGIC0'].update(build_aggregate[new_parent_id].trait_logic.button_text)

    layout_id = 1
    if build_aggregate[new_parent_id].child_id_list:

        for id in build_aggregate[new_parent_id].child_id_list:
            window['BUILD_ID' + str(layout_id)].update(value=build_aggregate[id].trait_id)
            window['TRAIT_NAME' + str(layout_id)].update(value=build_aggregate[id].trait_name)
            window['TRAIT_TABLE' + str(layout_id)].update(value=build_aggregate[id].rand_table)
            window['TRAIT_UNIT' + str(layout_id)].update(value=build_aggregate[id].trait_unit)
            window['is_primary' + str(layout_id)].update(value=build_aggregate[id].primary_trait)
            window['EXPAND_TRAIT' + str(layout_id)].update(disabled=False)
            window['LOGIC' + str(layout_id)].update(build_aggregate[id].trait_logic.button_text, disabled=False)

            layout_id += 1

    unique_build_id = int(next(reversed(build_aggregate))) + 1
    for empty_id in range(layout_id, 26):
        window['BUILD_ID' + str(empty_id)].update(value=unique_build_id)
        unique_build_id += 1
        window['TRAIT_NAME' + str(empty_id)].update(value="")
        window['TRAIT_TABLE' + str(empty_id)].update(value="")
        window['TRAIT_UNIT' + str(empty_id)].update(value="")
        window['is_primary' + str(empty_id)].update(value=False)
        window['EXPAND_TRAIT' + str(empty_id)].update(disabled=True)
        window['LOGIC' + str(empty_id)].update("LOGIC", disabled=True)

def refresh_agg_list(build_aggregate):
    global aggregate_list
    aggregate_list = []
    for trait in build_aggregate:
        aggregate_list.append(build_aggregate[trait].get_trait_info())


def main():
    global unique_build_id
    global unique_game_id
    global build_trait_parent_id
    global game_trait_parent_id
    #global build_aggregate
    build_aggregate = OrderedDict()
    global game_aggregate
    game_aggregate = OrderedDict()

    global list_of
    global aggregate_list

    case_test = ['']*6
    case_roll = ['']*6
    print (case_test)
    print(case_roll)

    last_case = ['']*6
    print(last_case)


    last_if = ""

    build_aggregate["0"] = Build_Trait("0")

    print(sg.Window.get_screen_size())

    global screen_w, screen_h
    # screen_w, screen_h = sg.Window.get_screen_size()  #Was using to size window based on Monitor size.
    # screen_w = int(screen_w*.45)
    # screen_h = int(screen_h*.8)

    # Static Window Size (Was trying to hit 1080P resolution, but couldn't fit all the lines.)
    screen_w = 1920
    screen_h = 1180

    # For possible future menu implementation
    menu_def = [['&File', ['E&xit']],
                ['&Help', ['&About']]]

    journal_layout = [[sg.Text('JOURNAL WINDOW')]]
    traits_layout = [[sg.Column(game_layout(), size=(int(screen_w / 2), int(screen_h * .8)), pad=(0, 0))]]
    right_layout = [[sg.Column(build_layout(build_aggregate), size=(int(screen_w / 2), int(screen_h * .8)),
                               background_color='blue', pad=(0, 0))],
                    [sg.Column(journal_layout, size=(int(screen_w / 2), int(screen_h * .2)), background_color='green',
                               pad=(0, 0))]]
    layout = [[sg.Column(traits_layout, size=(int(screen_w / 2), screen_h)),
               sg.Column(right_layout, size=(int(screen_w / 2), screen_h))]]

    window = sg.Window('Partonomy', layout, size=(screen_w, screen_h))

    do_this_once = True

    LOGIC_window_active = False
    while True:
        event, values = window.read(timeout=100)
        # Anything in this if clause will happen one time, after the UI is initialized.
        if do_this_once:
            #initialize_test_agg(build_aggregate)
            refresh_agg_list(build_aggregate)

            # for key in build_aggregate:
            #     print(build_aggregate[key])
            event = "EXPAND_TRAIT0"
            do_this_once = False


        if event in (None, 'Exit'):
            print("[LOG] Clicked Exit!")
            break

        if not LOGIC_window_active and event == 'LOAD_BUILD':
            filepath = sg.popup_get_file("Please select file",
                file_types=(("Partonomy Files", "*.dat"), ("All Files", "*.*"))
            )
            build_aggregate=pickle.load(open(filepath,"rb"))
            build_trait_parent_id="0"
            update_ui(window,build_trait_parent_id,build_aggregate)

        if not LOGIC_window_active and event[:5] == "LOGIC":  # LOOK AT LEFT 5 CHARACTERS of EVENT TO DETERMINE IF ANY LOGIC BUTTON HAS BEEN PRESSED
            LOGIC_window_active = True
            load_existing_logic = True
            logic_window = logic_tabbed_window()
            logic_key = event[5:]
            print("[LOG] User Entered LOGIC #" + logic_key + " menu: ")

        elif not LOGIC_window_active and event == 'SAVE_BUILD':
            filepath = sg.popup_get_file("Please select file", save_as=True,
                file_types=(("Partonomy Files", "*.dat"), ("All Files", "*.*"))
            )
            print("SAVING")
            for key in range(0, 26):
                if values[('TRAIT_TABLE' + str(key))] != "":
                    t_table = values[('TRAIT_TABLE' + str(key))]
                    t_name = values[('TRAIT_NAME' + str(key))]
                    if t_name == "":
                        window['TRAIT_NAME' + str(key)].update(value=t_table)
                        t_name = t_table

                    build_id = values[('BUILD_ID' + str(key))]
                    if key == 0:
                        trait_parent_id = build_aggregate[values['BUILD_ID' + str(key)]].build_parent_id
                    else:
                        trait_parent_id = build_trait_parent_id

                    temp_children_id = []
                    if (build_id in build_aggregate.keys()):
                        temp_children_id = build_aggregate[values['BUILD_ID' + str(key)]].child_id_list

                    if (values['TRAIT_TABLE' + str(key)] in list_of.keys()):
                        if (values['BUILD_ID'+str(key)] in build_aggregate.keys()) and build_aggregate[build_id].trait_logic.type != "NoLogic":
                            print("Current Build_agg has logic")
                            temp_logic=build_aggregate[build_id].trait_logic
                            print(temp_logic)
                        else:
                            print("No existing logic")
                            temp_logic=Build_Logic()
                            print(temp_logic)

                        build_aggregate[build_id] = Build_Trait(build_id,
                                                                trait_name=t_name,
                                                                build_parent_id=trait_parent_id,
                                                                rand_table=values['TRAIT_TABLE' + str(key)],
                                                                trait_unit=values['TRAIT_UNIT' + str(key)],
                                                                primary_trait=values['is_primary' + str(key)],
                                                                trait_logic=temp_logic)
                        build_aggregate[build_id].child_id_list = temp_children_id

                        if key != 0:
                            window['EXPAND_TRAIT' + str(key)].update(disabled=False)
                            window['LOGIC' + str(key)].update(disabled=False)

                            if build_id != "0" and not (build_aggregate[build_id].trait_id in build_aggregate[build_trait_parent_id].child_id_list):
                                build_aggregate[build_trait_parent_id].child_id_list.append(str(build_id))
                    else:
                        print("'TRAIT_TABLE-'", str(key), " =", values['TRAIT_TABLE' + str(key)], ", NOT IN LIST OF TABLES - TELL THE PLAYER!")
                else:
                    if values['TRAIT_NAME' + str(key)] != "":
                        print("TRAIT ", str(key), ": NOT SAVED --- TRAIT_NAME =", values['TRAIT_NAME' + str(key)], ", NO TRAIT SELECTED - TELL THE PLAYER!")

                pickle.dump(build_aggregate, open(filepath,"wb"), protocol=pickle.HIGHEST_PROTOCOL)

        elif not LOGIC_window_active and event == 'GO_TO_PARENT':
            if build_trait_parent_id != "0":
                build_trait_parent_id = str(build_aggregate[build_trait_parent_id].build_parent_id)
                update_ui(window, build_trait_parent_id,build_aggregate)
            else:
                print("Can't Go to Parent, Wouldn't be prudent! (You are in the Prime Trait)")

        elif not LOGIC_window_active and event[:12] == 'EXPAND_TRAIT':
            layout_id = event[12:]
            build_trait_parent_id = values[('BUILD_ID' + layout_id)]
            update_ui(window, build_trait_parent_id,build_aggregate)

        elif not LOGIC_window_active and event == 'GENERATE_AGGREGATE':
            # for k in build_aggregate:
            #     print(build_aggregate[k])

            unique_game_id = 0
            game_aggregate = OrderedDict()

            build_aggregate["0"].generate(build_aggregate)
            refresh_agg_list(build_aggregate)

            for k in game_aggregate:
                print(game_aggregate[k])

            window['GAME_ID0'].update(value=game_trait_parent_id)
            window['GAME_PROFICIENCY_RESULT0'].update(value=game_aggregate[game_trait_parent_id].proficiency_result)
            window['GAME_TRAIT_NAME0'].update(value=game_aggregate[game_trait_parent_id].trait_name)
            window['GAME_TRAIT_RESULT0'].update(
                value=game_aggregate[game_trait_parent_id].trait_result + " " + game_aggregate[
                    game_trait_parent_id].trait_unit)

            if game_aggregate[game_trait_parent_id].child_id_list:
                layout_id = 1

                for id in game_aggregate[game_trait_parent_id].child_id_list:
                    window['GAME_ID' + str(layout_id)].update(value=game_aggregate[id].trait_id)
                    window['GAME_PROFICIENCY_RESULT' + str(layout_id)].update(
                        value=game_aggregate[id].proficiency_result)
                    window['GAME_TRAIT_NAME' + str(layout_id)].update(value=game_aggregate[id].trait_name)
                    window['GAME_TRAIT_RESULT' + str(layout_id)].update(
                        value=game_aggregate[id].trait_result + " " + game_aggregate[id].trait_unit)
                    window['EXPAND_GAME_TRAIT' + str(layout_id)].update(disabled=False)
                    layout_id += 1

                for empty_id in range(layout_id, 26):
                    window['GAME_ID' + str(empty_id)].update(value=unique_build_id)
                    unique_build_id += 1
                    window['GAME_PROFICIENCY_RESULT' + str(empty_id)].update(value="")
                    window['GAME_TRAIT_NAME' + str(empty_id)].update(value="")
                    window['GAME_TRAIT_RESULT' + str(empty_id)].update(value="")
                    window['EXPAND_GAME_TRAIT' + str(empty_id)].update(disabled=True)

        elif not LOGIC_window_active and event == 'GO_TO_GAME_PARENT':
            if game_trait_parent_id != game_aggregate[game_trait_parent_id].game_parent_id:
                game_trait_parent_id = int(game_aggregate[game_trait_parent_id].game_parent_id)
                window['GAME_ID0'].update(value=game_trait_parent_id)
                window['GAME_PROFICIENCY_RESULT0'].update(
                    value=game_aggregate[game_trait_parent_id].proficiency_result)
                window['GAME_TRAIT_NAME0'].update(value=game_aggregate[game_trait_parent_id].trait_name)
                window['GAME_TRAIT_RESULT0'].update(
                    value=game_aggregate[game_trait_parent_id].trait_result + " " + game_aggregate[id].trait_unit)

                if game_aggregate[game_trait_parent_id].child_id_list:
                    layout_id = 1

                    for this_id in game_aggregate[game_trait_parent_id].child_id_list:
                        window['GAME_ID' + str(layout_id)].update(value=game_aggregate[this_id].trait_id)
                        window['GAME_PROFICIENCY_RESULT' + str(layout_id)].update(
                            value=game_aggregate[this_id].proficiency_result)
                        window['GAME_TRAIT_NAME' + str(layout_id)].update(value=game_aggregate[this_id].trait_name)
                        window['GAME_TRAIT_RESULT' + str(layout_id)].update(
                            value=game_aggregate[this_id].trait_result + " " + game_aggregate[this_id].trait_unit)
                        window['EXPAND_GAME_TRAIT' + str(layout_id)].update(disabled=False)
                        layout_id += 1

                    for empty_id in range(layout_id, 26):
                        window['GAME_ID' + str(empty_id)].update(value=unique_build_id)
                        unique_build_id += 1
                        window['GAME_PROFICIENCY_RESULT' + str(empty_id)].update(value="")
                        window['GAME_TRAIT_NAME' + str(empty_id)].update(value="")
                        window['GAME_TRAIT_RESULT' + str(empty_id)].update(value="")
                        window['EXPAND_GAME_TRAIT' + str(empty_id)].update(disabled=True)

                else:
                    for empty_id in range(1, 26):
                        window['BUILD_ID' + str(empty_id)].update(value=unique_build_id)
                        unique_build_id += 1
                        window['TRAIT_NAME' + str(empty_id)].update(value="")
                        window['TRAIT_TABLE' + str(empty_id)].update(value="")
                        window['TRAIT_UNIT' + str(empty_id)].update(value="")
                        window['is_primary' + str(empty_id)].update(value=False)
                        window['EXPAND_TRAIT' + str(empty_id)].update(disabled=True)

        elif not LOGIC_window_active and event[:17] == 'EXPAND_GAME_TRAIT':
            layout_id = event[17:]
            game_trait_parent_id = int(values[('GAME_ID' + layout_id)])
            window['GAME_ID0'].update(value=game_trait_parent_id)
            window['GAME_PROFICIENCY_RESULT0'].update(value=game_aggregate[game_trait_parent_id].proficiency_result)
            window['GAME_TRAIT_NAME0'].update(value=game_aggregate[game_trait_parent_id].trait_name)
            window['GAME_TRAIT_RESULT0'].update(
                value=game_aggregate[game_trait_parent_id].trait_result + " " + game_aggregate[
                    game_trait_parent_id].trait_unit)

            if game_aggregate[game_trait_parent_id].child_id_list:
                layout_id = 1
                for this_id in game_aggregate[game_trait_parent_id].child_id_list:
                    window['GAME_ID' + str(layout_id)].update(value=game_aggregate[this_id].trait_id)
                    window['GAME_PROFICIENCY_RESULT' + str(layout_id)].update(
                        value=game_aggregate[this_id].proficiency_result)
                    window['GAME_TRAIT_NAME' + str(layout_id)].update(value=game_aggregate[this_id].trait_name)
                    window['GAME_TRAIT_RESULT' + str(layout_id)].update(
                        value=game_aggregate[this_id].trait_result + " " + game_aggregate[this_id].trait_unit)
                    window['EXPAND_GAME_TRAIT' + str(layout_id)].update(disabled=False)
                    layout_id += 1

                for empty_id in range(layout_id, 26):
                    window['GAME_ID' + str(empty_id)].update(value=unique_build_id)
                    unique_build_id += 1
                    window['GAME_PROFICIENCY_RESULT' + str(empty_id)].update(value="")
                    window['GAME_TRAIT_NAME' + str(empty_id)].update(value="")
                    window['GAME_TRAIT_RESULT' + str(empty_id)].update(value="")
                    window['EXPAND_GAME_TRAIT' + str(empty_id)].update(disabled=True)
            else:
                for empty_id in range(1, 26):
                    window['GAME_ID' + str(empty_id)].update(value=unique_build_id)
                    unique_build_id += 1
                    window['GAME_PROFICIENCY_RESULT' + str(empty_id)].update(value="")
                    window['GAME_TRAIT_NAME' + str(empty_id)].update(value="")
                    window['GAME_TRAIT_RESULT' + str(empty_id)].update(value="")
                    window['EXPAND_GAME_TRAIT' + str(empty_id)].update(disabled=True)


        if LOGIC_window_active:
            close_logic_window=False
            logic_event, logic_values = logic_window.read(timeout=100)
            # print("logic_event=",logic_event)
            # for value in logic_values:
            #     print("logic_values[",value,"]=", logic_values[value])
            if logic_event == None:
                close_logic_window = True
            else:
                #print("logic_key=",logic_key)
                #print("values['BUILD_ID'+str(logic_key)]=",values['BUILD_ID'+str(logic_key)])
                if (values['BUILD_ID'+str(logic_key)] in build_aggregate.keys()) and build_aggregate[values['BUILD_ID'+str(logic_key)]].trait_logic.type != "NoLogic" and load_existing_logic:
                    active_logic=build_aggregate[values['BUILD_ID'+str(logic_key)]].trait_logic
                    if active_logic.type == "If-Then":
                        logic_window['if_else'].update(value=active_logic.if_else)
                        logic_window['if_test'].update(value=active_logic.if_test)
                        logic_window['if_condition'].update(value=active_logic.if_condition)
                        logic_window['if_test_string'].update(value=active_logic.if_test_string)
                        logic_window['if_true_trait'].update(value=active_logic.if_true_trait)
                        logic_window['if_else_trait'].update(value=active_logic.if_else_trait)
                        logic_window.Element('If-Then').Select()
                        last_if=""
                    elif active_logic.type == "Quantity":
                        logic_window['quantity_count'].update(value=active_logic.quantity_count)
                        logic_window.Element('Quantity').Select()
                    elif active_logic.type == "Chance":
                        logic_window['chance_percent'].update(value=active_logic.chance_percent)
                        logic_window.Element('Chance').Select()
                    elif active_logic.type == "Case":
                        for i in range (1,6):
                            if i == 1:
                                logic_window['case_1'].update(value=active_logic.case_case)
                                last_case[i]=(logic_values['case_1'])
                            logic_window['case_test_'+str(i)].update(value=active_logic.case_test[i])
                            logic_window['case_roll_'+str(i)].update(value=active_logic.case_roll[i])

                        logic_window.Element('Case').Select()
                    load_existing_logic = False


                if logic_values["Logic_Tabs"] == "Case":
                    for i in range(1,6):
                        if (logic_values['case_1']) and (logic_values['case_1'] != last_case[i]):
                            table=build_aggregate[logic_values['case_1'].split(":")[0]].rand_table
                            last_case[i] = logic_values['case_1']
                            if table in [*list_of]:
                                list_of_elements = list_of[table]
                                logic_window['case_test_'+ str(i)].update(values=list_of_elements, value=logic_values['case_test_'+str(i)])
                elif logic_values["Logic_Tabs"] == "If-Then":
                    if (logic_values['if_test']) and (logic_values['if_test'] != last_if):
                        table = build_aggregate[logic_values['if_test'].split(":")[0]].rand_table
                        last_if = (logic_values['if_test'])
                        if table in [*list_of]:
                            list_of_elements = list_of[table]
                            logic_window['if_test_string'].update(values=list_of_elements, value = logic_values['if_test_string'])

                if logic_event == 'Logic_cancel':
                    close_logic_window = True

                elif logic_event == 'Logic_ok':
                    print("LOGIC CONFIRMED")
                    temp_logic = Build_Logic()
                    build_id = values["BUILD_ID"+str(logic_key)]

                    if logic_values["Logic_Tabs"] == "NoLogic":
                        temp_logic = Build_Logic(type="NoLogic")
                        build_aggregate[values["BUILD_ID" + str(logic_key)]].trait_logic = temp_logic
                        window['LOGIC' + str(logic_key)].Update(temp_logic.button_text)
                        close_logic_window = True
                    elif logic_values["Logic_Tabs"] == "If-Then":
                        print("If-Then OK")
                        if logic_values["if_condition"] !="" and logic_values["if_test"] != "" and logic_values["if_test_string"] !="" and logic_values["if_true_trait"] != "" and logic_values["if_else_trait"] != "":
                            temp_logic = Build_Logic(type="If-Then", if_else = logic_values["if_else"], if_condition=logic_values["if_condition"],
                                                     if_test=logic_values["if_test"],if_test_string=logic_values["if_test_string"],
                                                     if_true_trait = logic_values["if_true_trait"], if_else_trait = logic_values["if_else_trait"])
                            build_aggregate[values["BUILD_ID" + str(logic_key)]].trait_logic = temp_logic
                            window['LOGIC' + str(logic_key)].Update("IF THEN")
                            close_logic_window = True
                        else:
                            print("Entry is OUT OF BOUNDS!")
                    elif logic_values["Logic_Tabs"] == "Quantity":
                        print("Quantity OK")
                        if (logic_values["quantity_count"]).isdigit() and int(logic_values["quantity_count"]) > 1 and int(logic_values["quantity_count"]) < 10:
                            temp_logic = Build_Logic(type="Quantity",quantity_count=int(logic_values['quantity_count']))
                            build_aggregate[values["BUILD_ID" + str(logic_key)]].trait_logic = temp_logic
                            window['LOGIC' + str(logic_key)].Update("QUANTITY"+logic_values['quantity_count'])
                            close_logic_window = True
                        else:
                            print("Entry is OUT OF BOUNDS!")
                    elif logic_values["Logic_Tabs"] == "Chance":
                        print("Chance OK")
                        if (logic_values["chance_percent"]).isdigit() and int(logic_values["chance_percent"]) > 0 and int(
                                logic_values["chance_percent"]) < 100:
                            temp_logic = Build_Logic(type="Chance",chance_percent=int(logic_values['chance_percent']))
                            build_aggregate[values["BUILD_ID" + str(logic_key)]].trait_logic = temp_logic
                            window['LOGIC' + str(logic_key)].Update("CHANCE"+logic_values['chance_percent'])
                            close_logic_window = True

                        else:
                            print("Entry is OUT OF BOUNDS!")
                    elif logic_values["Logic_Tabs"] == "Case":
                        print("Case OK")
                        last_case = ["" for i in range(0,6)]
                        case_case=""
                        if logic_values['case_1'] != "" and logic_values['case_test_1'] != "" and logic_values['case_roll_1'] != "":
                            case_case=logic_values['case_1']
                            case_test[1]=logic_values['case_test_1']
                            case_roll[1]=logic_values['case_roll_1']
                        if logic_values['case_1'] != "" and logic_values['case_test_2'] != "" and logic_values['case_roll_2'] != "":
                            case_case=logic_values['case_1']
                            case_test[2]=logic_values['case_test_2']
                            case_roll[2]=logic_values['case_roll_2']
                        if logic_values['case_1'] != "" and logic_values['case_test_3'] != "" and logic_values['case_roll_3'] != "":
                            case_case = logic_values['case_1']
                            case_test[3] = logic_values['case_test_3']
                            case_roll[3] = logic_values['case_roll_3']
                        if logic_values['case_1'] != "" and logic_values['case_test_4'] != "" and logic_values['case_roll_4'] != "":
                            case_case = logic_values['case_1']
                            case_test[4] = logic_values['case_test_4']
                            case_roll[4] = logic_values['case_roll_4']
                        if logic_values['case_1'] != "" and logic_values['case_test_5'] != "" and logic_values['case_roll_5'] != "":
                            case_case = logic_values['case_1']
                            case_test[5] = logic_values['case_test_5']
                            case_roll[5] = logic_values['case_roll_5']
                        if case_case != "":
                            temp_logic = Build_Logic(type="Case",case_case=case_case, case_test=case_test,case_roll=case_roll)
                            build_aggregate[values["BUILD_ID" + str(logic_key)]].trait_logic = temp_logic
                            window['LOGIC' + str(logic_key)].Update("CASE")
                            close_logic_window = True
                        else:
                            print("Entry is OUT OF BOUNDS!")

                    elif logic_values["Logic_Tabs"] == "Exploding":
                        print("Exploding OK")
                    elif logic_values["Logic_Tabs"] == "Inherit":
                        print("Inherit OK")

            if close_logic_window:
                LOGIC_window_active = False
                logic_window.close()

    window.close()
    exit(0)





import_random_lists_from_file()  # TURNED OFF FOR TESTING UI

if __name__ == '__main__':
    main()
